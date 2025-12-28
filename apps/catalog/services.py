import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from io import BytesIO

class ExcelGenerator:
    """
    Generic Advanced Excel Generator.
    - Uses 'write_only=True' for O(1) memory usage with large datasets.
    - Supports declarative column definitions.
    """
    
    def __init__(self, title="Export", creator="OwlStore System"):
        self.workbook = openpyxl.Workbook(write_only=True)
        self.worksheet = self.workbook.create_sheet(title=title)
        self.workbook.properties.creator = creator
        
        # Styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo-600
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        
    def generate(self, queryset, columns):
        """
        Generates Excel file stream.
        :param queryset: Django QuerySet (will be iterated)
        :param columns: List of dicts/tuples [('Header', 'field_name'), ...]
        """
        # 1. Write Header
        header_row = []
        for col_def in columns:
            cell = openpyxl.cell.WriteOnlyCell(self.worksheet, value=col_def['header'])
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            header_row.append(cell)
        self.worksheet.append(header_row)

        # 2. Write Data (Iterator for memory efficiency)
        for obj in queryset.iterator():
            row = []
            for col_def in columns:
                value = self._get_value(obj, col_def['field'])
                
                # Optional: Formatter
                if 'formatter' in col_def and callable(col_def['formatter']):
                    value = col_def['formatter'](value)
                
                row.append(value)
            self.worksheet.append(row)

        # 3. Save to Bytes
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output

    def _get_value(self, obj, field):
        """Helper to get value from object, supporting dot notation (category.name)"""
        try:
            if callable(field):
                return field(obj)
            
            # Dot notation support
            value = obj
            for attr in field.split('.'):
                value = getattr(value, attr, '')
                if value is None:
                    break
            return value
        except Exception:
            return ''


class ProductExportService:
    @staticmethod
    def export_to_excel(queryset):
        generator = ExcelGenerator(title="Product List")
        
        # Declarative Column Configuration
        columns = [
            {'header': 'ID', 'field': 'id'},
            {'header': 'SKU', 'field': 'sku'},
            {'header': 'Product Name', 'field': 'name'},
            {'header': 'Category', 'field': 'category.name'},
            {'header': 'Brand', 'field': 'brand'},
            {'header': 'Price', 'field': 'price', 'formatter': lambda x: float(x) if x else 0},
            {'header': 'Sale Price', 'field': 'sale_price', 'formatter': lambda x: float(x) if x else 0},
            {'header': 'Stock', 'field': 'stock'},
            {'header': 'Active', 'field': 'is_active', 'formatter': lambda x: 'Yes' if x else 'No'},
            {'header': 'Featured', 'field': 'is_featured', 'formatter': lambda x: 'Yes' if x else 'No'},
            {'header': 'Created At', 'field': 'created_at', 'formatter': lambda x: x.strftime('%Y-%m-%d %H:%M') if x else ''},
        ]

        excel_file = generator.generate(queryset, columns)

        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="products_export.xlsx"'
        return response
